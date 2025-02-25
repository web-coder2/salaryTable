import openpyxl
import json

def find_next_week_start(sheet, start_row):
    """
    Finds the starting row of the next week's data by searching for a date.

    Args:
        sheet: The worksheet object.
        start_row: The row to start searching from.

    Returns:
        The row number of the next week's data, or None if not found.
    """
    row = start_row + 1  # Start searching from the next row
    while row <= sheet.max_row:
        date_cell = sheet.cell(row=row - 1, column=2).value  # Check first date column

        if date_cell is not None: # Found a non-empty date cell
            return row
        row += 1
    return None  # No next week found

def extract_data_from_excel(excel_file):
    """
    Extracts data from the specified Excel file and converts it into the specified JSON format,
    handling variable gaps between weeks.

    Args:
        excel_file (str): The path to the Excel file.

    Returns:
        str: A JSON string representing the extracted data.
    """

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active  # Assuming the relevant data is on the first sheet

    data = []
    week_start_row = 2  # Start row for the first week

    while week_start_row is not None: # Loop until no more weeks are found

        for day_offset in range(7):
            col_offset = day_offset * 2

            # Extract date
            date_cell = sheet.cell(row=week_start_row - 1, column=2 + col_offset).value
            date = str(date_cell) if date_cell else None

            # Extract robot value
            robot_cell = sheet.cell(row=week_start_row, column=2 + col_offset).value
            robot = robot_cell if robot_cell is not None else None

            # Extract Summa Hold values
            summa_hold_raznica_cell = sheet.cell(row=week_start_row + 1, column=2 + col_offset).value # Left cell
            summa_hold_itogo_cell = sheet.cell(row=week_start_row + 1, column=3 + col_offset).value  # Right Cell
            summa_hold_raznica = summa_hold_raznica_cell if summa_hold_raznica_cell is not None else None
            summa_hold_itogo = summa_hold_itogo_cell if summa_hold_itogo_cell is not None else None

            # Extract Okladniki value
            okladniki_cell = sheet.cell(row=week_start_row + 2, column=3 + col_offset).value
            okladniki = okladniki_cell if okladniki_cell is not None else None

            # Extract office value
            office_cell = sheet.cell(row=week_start_row + 3, column=3 + col_offset).value
            office = office_cell if office_cell is not None else None

            # Create the dictionary
            day_data = {
                "Дата": date,
                "Робот": robot,
                "Сумма_HOLD_Разница": summa_hold_raznica,
                "Сумма_HOLD_Итого": summa_hold_itogo,
                "Окладчики": okladniki,
                "Офис": office
            }
            data.append(day_data)

        # Find the start of the next week
        week_start_row = find_next_week_start(sheet, week_start_row + 10) # +4 to skip the current week's data

    return json.dumps(data, indent=4, ensure_ascii=False)

# Example usage (replace 'your_excel_file.xlsx' with the actual path)
excel_file = 'your_excel_file.xlsx'
json_data = extract_data_from_excel(excel_file)
print(json_data)

# To save to a JSON file
with open("output.json", "w", encoding="utf-8") as f:
    f.write(json_data)